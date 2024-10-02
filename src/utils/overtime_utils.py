import openpyxl

from openpyxl.styles.fonts import Font
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter

from config import MEAL_FEE


def check_overtime_pay(file_path: str) -> None:
    """
    초과근무 수당 체크 프로세스
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[wb.sheetnames[0]]

    row_len = len(ws['A'])

    for i in range(row_len-1, 0, -1):
        if ws['U'][i].value == 'X' or ws['F'][i].value in ['김수현', '이상수']:
            ws.delete_rows(i+1, 1)
        else:
            if ws['I'][i].value in ['수요일', '토요일', '일요일']:
                if input(ws['F'][i].value + " | " + ws['H'][i].value.strftime("%Y-%m-%d") + " | " + ws['I'][i].value + " | 사전 보고 확인? :").upper() == 'N':
                    ws.delete_rows(i+1, 1)

    wb.save(file_path)


def overtimeCnt(filename):
    overtimeNameCnt = {}

    wb = openpyxl.load_workbook(filename)
    ws = wb[wb.sheetnames[0]]

    dateCnt = {}
    maxCnt = 0

    for row_data in ws.iter_rows(2):
        if row_data[7].value.strftime("%Y-%m-%d") in dateCnt:
            dateCnt[row_data[7].value.strftime("%Y-%m-%d")] += 1
        else:
            dateCnt[row_data[7].value.strftime("%Y-%m-%d")] = 1

        if row_data[5].value in overtimeNameCnt:
            overtimeNameCnt[row_data[5].value] += 1
        else:
            overtimeNameCnt[row_data[5].value] = 1

        maxCnt += 1

    dateCnt = sorted(dateCnt.items())

    ws2 = wb.create_sheet("매식비 통계", 0)
    ColumnDimension(ws2, bestFit=True)

    # 데이터 채우기
    ws2['B2'] = "초과근무일자"
    ws2['C2'] = '인원'
    ws2['D2'] = '단가'
    ws2['E2'] = '금액'
    ws2['F2'] = '비고'

    for i in range(0, len(dateCnt)):
        j = str(i+3)
        (date, cnt) = dateCnt[i]
        ws2['B'+j] = date
        ws2['C'+j] = cnt
        ws2['D'+j] = MEAL_FEE
        ws2['E'+j] = cnt*MEAL_FEE

    lastRow = str(len(dateCnt)+3)
    ws2['B'+lastRow] = '합계'
    ws2['C'+lastRow] = maxCnt
    ws2['D'+lastRow] = ''
    ws2['E'+lastRow] = maxCnt*MEAL_FEE

    # Style
    font_format = Font(size=11, name='맑은 고딕')
    font_format_bold = Font(size=11, name='맑은 고딕', bold=True)
    border_format = Side(border_style="thin")
    align_format_center = Alignment(horizontal="center", vertical="center")
    align_format_vertical = Alignment(vertical="center")
    fill_style = PatternFill(start_color="00C0C0C0",
                             end_color="00C0C0C0", patternType="solid")

    for i in ['B', 'C', 'D', 'E', 'F']:
        for j in range(0, len(dateCnt)+2):
            k = str(j+2)
            ws2[i+k].font = font_format
            ws2[i+k].border = Border(top=border_format, bottom=border_format,
                                     left=border_format, right=border_format)
            if i in ['B', 'C'] or j == 0:
                ws2[i+k].alignment = align_format_center
            else:
                ws2[i+k].alignment = align_format_vertical

            if j in [0, len(dateCnt)+1]:
                ws2[i+k].font = font_format_bold
                ws2[i+k].fill = fill_style

            if i in ['D', 'E'] and j > 0:
                ws2[i+k].number_format = '#,##0'

    for column_cells in ws2.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))

        if new_column_length > 0:
            ws2.column_dimensions[new_column_letter].width = new_column_length*1.7

    wb.save(filename)

    return overtimeNameCnt


def officialDataMaker(filename, overtimeNameCnt):
    wb = openpyxl.load_workbook(filename)
    ws = wb[wb.sheetnames[0]]

    data = {}

    row_len = len(ws['A'])
    for i in range(2, row_len-1):
        data[ws['I'][i].value] = [
            int(ws['K'][i].value.split(':')[0]), int(ws['AC'][i].value)]

    for name in ['윤인자', '이종선', '홍성민', '황미연', '강동욱', '우미인']:
        try:
            overtimeNameCnt[name]
        except KeyError:
            overtimeNameCnt[name] = 0

        print("%s | %2d | %2d | %2d" %
              (name, data[name][0], data[name][1], overtimeNameCnt[name]))
